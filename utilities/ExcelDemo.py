import openpyxl


class ExcelDemo:
    def getdataFromExcel(self):
        # Open the Excel file
        workbook = openpyxl.load_workbook("D:\\Frameworks\\PythonExcelDemo.xlsx")

        # Select the sheet by name
        sheet = workbook["LoginPage"]

        # Get data from Excel cells
        username12 = sheet.cell(row=2, column=1).value
        password13 = sheet.cell(row=2, column=2).value

        print(f"Username: {username12}, Password: {password13}")

        # Close the Excel file
        workbook.close()

# run the method
ExcelDemo().getdataFromExcel()
