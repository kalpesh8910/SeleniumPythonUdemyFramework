import time

import openpyxl
import pytest

from TestData.CreateChannelPageData import CreateChannelPageData
from TestData.LoginPageData import LoginPageData
from pageObjects.CreateChannel import CreateChannel
from pageObjects.LoginPage import LoginPage
from utilities.BaseClass import BaseClass


class TestOne(BaseClass):

    def test_e2e(self, setup, getData, getDataForChannel):

        # Open the Excel file
        workbook = openpyxl.load_workbook("D:\\Frameworks\\PythonExcelDemo.xlsx")

        # Select the sheet by name
        sheet1 = workbook["LoginPage"]
        sheet2 = workbook["CreateChannel"]

        log = self.getLogger()

        LOGINPAGE = LoginPage(self.driver)
        log.info("Login page object created.")

        CREATECHANNEL = CreateChannel(self.driver)
        log.info("Create channel object created.")

        # Enter username
        LOGINPAGE.enterUsername().send_keys(sheet1.cell(row=2, column=1).value)
        #  LOGINPAGE.enterUsername().send_keys(getData["username"])
        log.info("username is:-" + getData["username"])
        time.sleep(2)

        # Enter password
        LOGINPAGE.enterPassword().send_keys(sheet1.cell(row=2, column=2).value)
        #  LOGINPAGE.enterPassword().send_keys(getData["password"])
        log.info("password is:-" + getData["password"])
        time.sleep(2)

        # Click on let's go button
        LOGINPAGE.clickLoginButton().click()
        log.info("Clicked on let's go button")
        time.sleep(7)

        # Click on +New channel button
        CREATECHANNEL.createLink().click()
        log.info("Click on +New channel button")
        time.sleep(5)

        # Enter Recipient name
        CREATECHANNEL.enterChannelName().send_keys(sheet2.cell(row=2, column=1).value)
        #  CREATECHANNEL.enterChannelName().send_keys(getDataForChannel["RecipientName"])
        log.info("Recipient name is:-" + getDataForChannel["RecipientName"])
        time.sleep(5)

        # Enter company name
        CREATECHANNEL.enterCompanyName().send_keys(sheet2.cell(row=2, column=2).value)
        #  CREATECHANNEL.enterCompanyName().send_keys(getDataForChannel["CompanyName"])
        log.info("company name is:-" + getDataForChannel["CompanyName"])
        time.sleep(2)

        # Enter email address
        CREATECHANNEL.enterEmailAddress().send_keys(sheet2.cell(row=2, column=3).value)
        #  CREATECHANNEL.enterEmailAddress().send_keys(getDataForChannel["EmailAddress"])
        log.info("Email address is:-" + getDataForChannel["EmailAddress"])
        time.sleep(2)

        # Enter tagline
        CREATECHANNEL.enterTagLine().send_keys(sheet2.cell(row=2, column=4).value)
        #  CREATECHANNEL.enterTagLine().send_keys(getDataForChannel["Tagline"])
        log.info("tag line is:-" + getDataForChannel["Tagline"])
        time.sleep(3)

        # Click on create button on popup screen
        CREATECHANNEL.clickOnCreateButton().click()
        log.info("Click on create button")
        time.sleep(7)

    #    msg = CREATECHANNEL.verifyMsg().text
     #   print(msg)

     #   assert ("nik karlos11111" in msg)

    @pytest.fixture(params=LoginPageData.test_HomePage_data)
    def getData(self, request):
        return request.param

    @pytest.fixture(params=CreateChannelPageData.test_CreateChannel_data)
    def getDataForChannel(self, request):
        return request.param
